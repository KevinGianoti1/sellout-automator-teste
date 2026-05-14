import os
from datetime import datetime, timedelta
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== HELPERS DE ESTILO =====

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)

def _border() -> Border:
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _header_row(ws, linha: int, headers: List[str], bg="1F4E79"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=linha, column=col, value=h)
        c.fill = _fill(bg)
        c.font = _font(bold=True, color="FFFFFF")
        c.border = _border()
        c.alignment = _center()

def _data_row(ws, linha: int, valores: List, bg=None):
    alt_bg = "F2F2F2" if linha % 2 == 0 else "FFFFFF"
    fill = _fill(bg or alt_bg)
    for col, v in enumerate(valores, 1):
        c = ws.cell(row=linha, column=col, value=v)
        c.fill = fill
        c.border = _border()
        c.alignment = _center() if col > 1 else _left()

def _ajustar_colunas(ws, min_w=12, max_w=45):
    for col in ws.columns:
        w = min_w
        for cell in col:
            if cell.value:
                w = max(w, min(len(str(cell.value)) + 3, max_w))
        ws.column_dimensions[col[0].column_letter].width = w

def _titulo(ws, texto: str, subtitulo: str = ""):
    ws['A1'] = texto
    ws['A1'].font = _font(bold=True, size=14, color="1F4E79")
    ws['A2'] = subtitulo or f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = _font(italic=True, size=10, color="666666")

def _kpi(ws, linha: int, label: str, valor, col=1, cor_valor="1F4E79"):
    ws.cell(row=linha, column=col, value=label).font = _font(bold=True, size=10, color="555555")
    c = ws.cell(row=linha, column=col + 1, value=valor)
    c.font = _font(bold=True, size=12, color=cor_valor)

def _moeda(v) -> str:
    try:
        return f"R$ {float(v):,.2f}"
    except (TypeError, ValueError):
        return "R$ 0,00"


# ===== RELATÓRIO 1: COMERCIAL OPERACIONAL =====

def gerar_relatorio_comercial_operacional(
    snapshot: Dict[str, Dict[str, int]],
    vendedores: List[Dict],
    finalizados: Dict,
    funis_comercial: Dict[str, str],
    filepath: str,
) -> None:
    """4 abas: Resumo, Por Vendedor, Em Andamento, Finalizados."""
    wb = Workbook()
    wb.remove(wb.active)

    # ---- ABA 1: RESUMO ----
    ws = wb.create_sheet("Resumo")
    _titulo(ws, "COMERCIAL OPERACIONAL - RESUMO")

    total_andamento = sum(v['em_andamento'] for v in vendedores)
    total_criadas = sum(v.get('criadas', v['em_andamento'] + v['ganhos_3m'] + v['perdidos_3m']) for v in vendedores)
    total_ganhos = sum(v['ganhos_3m'] for v in vendedores)
    total_perdidos = sum(v['perdidos_3m'] for v in vendedores)
    total_valor = sum(v['valor_ganhos_3m'] for v in vendedores)

    linha = 4
    ws.cell(row=linha, column=1, value="INDICADORES (ultimos 3 meses)").font = _font(bold=True, size=12, color="1F4E79")
    linha += 1

    kpis = [
        ("Criadas (andamento + ganhos + perdidos)", total_criadas, "1F4E79"),
        ("Em Andamento (estado atual)", total_andamento, "1F4E79"),
        ("Ganhos - ultimos 3 meses", total_ganhos, "006100"),
        ("Perdidos - ultimos 3 meses", total_perdidos, "C00000"),
        ("Valor Ganhos 3m", _moeda(total_valor), "006100"),
    ]
    for label, valor, cor in kpis:
        _kpi(ws, linha, label, valor, col=1, cor_valor=cor)
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="FINALIZADOS POR FUNIL (ultimos 3 meses)").font = _font(bold=True, size=12, color="1F4E79")
    linha += 1
    _header_row(ws, linha, ["Funil", "Ganhos", "Valor Ganhos (R$)", "Perdidos"])
    linha += 1
    for nome_funil, dados in finalizados.items():
        if nome_funil in funis_comercial:
            _data_row(ws, linha, [
                nome_funil,
                dados['won']['count'],
                _moeda(dados['won']['value']),
                dados['lost']['count'],
            ])
            linha += 1

    _ajustar_colunas(ws)

    # ---- ABA 2: POR VENDEDOR ----
    ws2 = wb.create_sheet("Por Vendedor")
    _titulo(ws2, "COMERCIAL - POR VENDEDOR")

    linha = 4
    _header_row(ws2, linha, ["Vendedor", "Criadas", "Em Andamento", "Ganhos 3m", "Perdidos 3m", "Valor Ganhos 3m (R$)"])
    linha += 1
    for v in vendedores:
        criadas = v.get('criadas', v['em_andamento'] + v['ganhos_3m'] + v['perdidos_3m'])
        _data_row(ws2, linha, [
            v['vendedor'],
            criadas,
            v['em_andamento'],
            v['ganhos_3m'],
            v['perdidos_3m'],
            _moeda(v['valor_ganhos_3m']),
        ])
        linha += 1

    _ajustar_colunas(ws2)

    # ---- ABA 3: EM ANDAMENTO (a partir da 2a etapa) ----
    ws3 = wb.create_sheet("Em Andamento")
    _titulo(ws3, "COMERCIAL - LEADS EM ANDAMENTO")
    ws3['A2'] = "Leads a partir da 2a etapa de cada funil comercial"
    ws3['A2'].font = _font(italic=True, size=10, color="666666")

    linha = 4
    _header_row(ws3, linha, ["Funil", "Estagio", "Quantidade"])
    linha += 1

    for funil_nome in funis_comercial:
        if funil_nome not in snapshot:
            continue
        stages = list(snapshot[funil_nome].items())
        for i, (stage_name, count) in enumerate(stages):
            if i == 0:
                continue  # pula primeira etapa
            _data_row(ws3, linha, [funil_nome, stage_name, count])
            linha += 1

    _ajustar_colunas(ws3)

    # ---- ABA 4: FINALIZADOS ----
    ws4 = wb.create_sheet("Finalizados")
    _titulo(ws4, "COMERCIAL - FINALIZADOS (ultimos 3 meses)")

    linha = 4
    _header_row(ws4, linha, ["Funil", "Status", "Quantidade", "Valor Total (R$)"])
    linha += 1

    for nome_funil, dados in finalizados.items():
        if nome_funil not in funis_comercial:
            continue
        c_won = ws4.cell(row=linha, column=1, value=nome_funil)
        c_won.border = _border(); c_won.alignment = _left()
        c_won.fill = _fill("FFFFFF" if linha % 2 == 0 else "F2F2F2")

        for col, v in enumerate(["Ganho", dados['won']['count'], _moeda(dados['won']['value'])], 2):
            c = ws4.cell(row=linha, column=col, value=v)
            c.border = _border(); c.alignment = _center()
            c.fill = _fill("E2EFDA")
            if col == 2:
                c.font = _font(color="006100", bold=True)
        linha += 1

        c_lost = ws4.cell(row=linha, column=1, value=nome_funil)
        c_lost.border = _border(); c_lost.alignment = _left()
        c_lost.fill = _fill("FFFFFF" if linha % 2 == 0 else "F2F2F2")
        for col, v in enumerate(["Perdido", dados['lost']['count'], "—"], 2):
            c = ws4.cell(row=linha, column=col, value=v)
            c.border = _border(); c.alignment = _center()
            c.fill = _fill("FFDCE1")
            if col == 2:
                c.font = _font(color="C00000", bold=True)
        linha += 1

    _ajustar_colunas(ws4)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"   Salvo: {filepath}")


# ===== RELATÓRIO 2: COMERCIAL CONVERSÃO =====

def gerar_relatorio_comercial_conversao(
    snapshot: Dict[str, Dict[str, int]],
    taxas: Dict[str, List[Dict]],
    funis_comercial: Dict[str, str],
    filepath: str,
) -> None:
    """2 abas: Taxa de Conversão, Gargalos."""
    wb = Workbook()
    wb.remove(wb.active)

    # ---- ABA 1: TAXA DE CONVERSÃO ----
    ws = wb.create_sheet("Taxa de Conversao")
    _titulo(ws, "COMERCIAL - TAXA DE CONVERSAO ENTRE ESTAGIOS")

    linha = 4
    _header_row(ws, linha, ["Funil", "Estagio", "Leads", "Proximo Estagio", "Leads Proximo", "Conversao (%)"])
    linha += 1

    for funil_nome in funis_comercial:
        if funil_nome not in taxas:
            continue
        for etapa in taxas[funil_nome]:
            taxa = etapa['taxa_pct']
            if taxa is None:
                bg = "D9D9D9"
            elif taxa >= 50:
                bg = "C6EFCE"  # verde
            elif taxa >= 20:
                bg = "FFEB9C"  # amarelo
            else:
                bg = "FFC7CE"  # vermelho

            valores = [
                funil_nome,
                etapa['estagio'],
                etapa['count'],
                etapa['proximo_estagio'] or "—",
                etapa['proximo_count'] if etapa['proximo_count'] is not None else "—",
                f"{taxa}%" if taxa is not None else "—",
            ]
            for col, v in enumerate(valores, 1):
                c = ws.cell(row=linha, column=col, value=v)
                c.fill = _fill(bg)
                c.border = _border()
                c.alignment = _left() if col == 1 else _center()
            linha += 1

    # Legenda
    linha += 1
    ws.cell(row=linha, column=1, value="Legenda de cores:").font = _font(bold=True)
    linha += 1
    for cor, texto in [("C6EFCE", "Verde: conversao >= 50%"), ("FFEB9C", "Amarelo: conversao entre 20% e 49%"), ("FFC7CE", "Vermelho: conversao < 20%")]:
        c = ws.cell(row=linha, column=1, value=texto)
        c.fill = _fill(cor)
        c.border = _border()
        linha += 1

    _ajustar_colunas(ws)

    # ---- ABA 2: GARGALOS ----
    ws2 = wb.create_sheet("Gargalos")
    _titulo(ws2, "COMERCIAL - GARGALOS DO FUNIL")
    ws2['A2'] = "Estagios com mais leads acumulados e menor taxa de saida"
    ws2['A2'].font = _font(italic=True, size=10, color="666666")

    gargalos = []
    for funil_nome in funis_comercial:
        if funil_nome not in taxas:
            continue
        for etapa in taxas[funil_nome]:
            if etapa['taxa_pct'] is not None and etapa['taxa_pct'] < 30 and etapa['count'] > 0:
                gargalos.append({
                    'funil': funil_nome,
                    'estagio': etapa['estagio'],
                    'leads': etapa['count'],
                    'taxa': etapa['taxa_pct'],
                })

    gargalos.sort(key=lambda x: (-x['leads'], x['taxa']))

    linha = 4
    _header_row(ws2, linha, ["Funil", "Estagio (Gargalo)", "Leads Acumulados", "Taxa de Saida (%)"])
    linha += 1

    if gargalos:
        for g in gargalos:
            _data_row(ws2, linha, [g['funil'], g['estagio'], g['leads'], f"{g['taxa']}%"])
            linha += 1
    else:
        ws2.cell(row=linha, column=1, value="Nenhum gargalo critico identificado (taxa < 30%)").font = _font(italic=True, color="666666")

    _ajustar_colunas(ws2)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"   Salvo: {filepath}")


# ===== RELATÓRIO 3: MARKETING =====

def gerar_relatorio_marketing(
    snapshot: Dict[str, Dict[str, int]],
    vendedores_mkt: List[Dict],
    finalizados_mkt: Dict,
    taxas_mkt: Dict[str, List[Dict]],
    funis_marketing: Dict[str, str],
    filepath: str,
) -> None:
    """4 abas: Resumo, Por Vendedor, Jornada no Funil, Finalizados."""
    wb = Workbook()
    wb.remove(wb.active)

    # ---- ABA 1: RESUMO ----
    ws = wb.create_sheet("Resumo")
    _titulo(ws, "MARKETING - FUNIL DE CLIENTES NOVOS")

    total_andamento = sum(v['em_andamento'] for v in vendedores_mkt)
    total_ganhos = sum(v['ganhos_3m'] for v in vendedores_mkt)
    total_perdidos = sum(v['perdidos_3m'] for v in vendedores_mkt)
    total_valor = sum(v['valor_ganhos_3m'] for v in vendedores_mkt)
    total_leads_funil = sum(sum(snapshot.get(f, {}).values()) for f in funis_marketing)

    linha = 4
    ws.cell(row=linha, column=1, value="INDICADORES (ultimos 3 meses)").font = _font(bold=True, size=12, color="1F4E79")
    linha += 1
    kpis = [
        ("Total no Funil (atual)", total_leads_funil, "1F4E79"),
        ("Leads em Andamento", total_andamento, "1F4E79"),
        ("Ganhos", total_ganhos, "006100"),
        ("Perdidos", total_perdidos, "C00000"),
        ("Valor Ganhos", _moeda(total_valor), "006100"),
    ]
    for label, valor, cor in kpis:
        _kpi(ws, linha, label, valor, cor_valor=cor)
        linha += 1

    _ajustar_colunas(ws)

    # ---- ABA 2: POR VENDEDOR ----
    ws2 = wb.create_sheet("Por Vendedor")
    _titulo(ws2, "MARKETING - POR VENDEDOR")

    linha = 4
    _header_row(ws2, linha, ["Vendedor", "Em Andamento", "Ganhos 3m", "Perdidos 3m", "Valor Ganhos 3m (R$)"])
    linha += 1
    for v in vendedores_mkt:
        _data_row(ws2, linha, [
            v['vendedor'],
            v['em_andamento'],
            v['ganhos_3m'],
            v['perdidos_3m'],
            _moeda(v['valor_ganhos_3m']),
        ])
        linha += 1

    _ajustar_colunas(ws2)

    # ---- ABA 3: JORNADA NO FUNIL ----
    ws3 = wb.create_sheet("Jornada no Funil")
    _titulo(ws3, "MARKETING - JORNADA NO FUNIL")

    linha = 4
    _header_row(ws3, linha, ["Funil", "Estagio", "Leads", "Proximo Estagio", "Leads Proximo", "Conversao (%)"])
    linha += 1

    for funil_nome in funis_marketing:
        if funil_nome not in taxas_mkt:
            continue
        for etapa in taxas_mkt[funil_nome]:
            taxa = etapa['taxa_pct']
            bg = "C6EFCE" if (taxa or 0) >= 50 else ("FFEB9C" if (taxa or 0) >= 20 else ("FFC7CE" if taxa is not None else "D9D9D9"))
            valores = [
                funil_nome,
                etapa['estagio'],
                etapa['count'],
                etapa['proximo_estagio'] or "—",
                etapa['proximo_count'] if etapa['proximo_count'] is not None else "—",
                f"{taxa}%" if taxa is not None else "—",
            ]
            for col, v in enumerate(valores, 1):
                c = ws3.cell(row=linha, column=col, value=v)
                c.fill = _fill(bg)
                c.border = _border()
                c.alignment = _left() if col == 1 else _center()
            linha += 1

    _ajustar_colunas(ws3)

    # ---- ABA 4: FINALIZADOS ----
    ws4 = wb.create_sheet("Finalizados")
    _titulo(ws4, "MARKETING - FINALIZADOS (ultimos 3 meses)")

    linha = 4
    _header_row(ws4, linha, ["Funil", "Status", "Quantidade", "Valor Total (R$)"])
    linha += 1

    for nome_funil, dados in finalizados_mkt.items():
        for col, v in enumerate([nome_funil, "Ganho", dados['won']['count'], _moeda(dados['won']['value'])], 1):
            c = ws4.cell(row=linha, column=col, value=v)
            c.fill = _fill("E2EFDA"); c.border = _border(); c.alignment = _left() if col == 1 else _center()
        linha += 1
        for col, v in enumerate([nome_funil, "Perdido", dados['lost']['count'], "—"], 1):
            c = ws4.cell(row=linha, column=col, value=v)
            c.fill = _fill("FFDCE1"); c.border = _border(); c.alignment = _left() if col == 1 else _center()
        linha += 1

    _ajustar_colunas(ws4)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"   Salvo: {filepath}")


# ===== EMAIL HTML RESUMO =====

def gerar_email_html_resumo(
    vendedores: List[Dict],
    finalizados: Dict,
    vendedores_mkt: List[Dict],
    finalizados_mkt: Dict,
) -> str:
    data_str = datetime.now().strftime("%d/%m/%Y")
    hora_str = datetime.now().strftime("%H:%M")

    total_criadas_com = sum(v.get('criadas', v['em_andamento'] + v['ganhos_3m'] + v['perdidos_3m']) for v in vendedores)
    total_andamento = sum(v['em_andamento'] for v in vendedores)
    total_ganhos_com = sum(v['ganhos_3m'] for v in vendedores)
    total_perdidos_com = sum(v['perdidos_3m'] for v in vendedores)
    total_valor_com = sum(v['valor_ganhos_3m'] for v in vendedores)

    total_andamento_mkt = sum(v['em_andamento'] for v in vendedores_mkt)
    total_ganhos_mkt = sum(v['ganhos_3m'] for v in vendedores_mkt)
    total_perdidos_mkt = sum(v['perdidos_3m'] for v in vendedores_mkt)
    total_valor_mkt = sum(v['valor_ganhos_3m'] for v in vendedores_mkt)

    todas_rows = "".join(
        f"<tr>"
        f"<td><b>{v['vendedor']}</b></td>"
        f"<td style='text-align:center;color:#00b4d8;font-weight:bold'>{v.get('criadas', v['em_andamento'] + v['ganhos_3m'] + v['perdidos_3m'])}</td>"
        f"<td style='text-align:center;color:#00b4d8'>{v['em_andamento']}</td>"
        f"<td style='text-align:center;color:#06d6a0'>{v['ganhos_3m'] + v['perdidos_3m']}</td>"
        f"</tr>"
        for v in vendedores if v['em_andamento'] > 0 or v['ganhos_3m'] > 0 or v['perdidos_3m'] > 0
    )

    periodo_inicio = (datetime.now().replace(day=1) - timedelta(days=60)).strftime("%d/%m/%y")
    periodo_fim = datetime.now().strftime("%d/%m/%y")

    return f"""<!DOCTYPE html>
<html>
<head><meta charset='UTF-8'><style>
body{{font-family:Arial,sans-serif;background:#1a1a2e;margin:0;padding:20px}}
.wrap{{max-width:700px;margin:0 auto;background:#16213e;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.4)}}
.header{{background:#0f3460;color:#fff;padding:24px 30px}}
.header h1{{margin:0;font-size:20px;color:#00b4d8}}
.header p{{margin:4px 0 0;font-size:12px;opacity:.7;color:#ccc}}
.body{{padding:24px 30px}}
.section-title{{font-size:13px;font-weight:bold;color:#00b4d8;border-bottom:1px solid #00b4d8;padding-bottom:6px;margin:20px 0 12px;display:flex;justify-content:space-between;align-items:center}}
.badge{{background:#0f3460;color:#00b4d8;font-size:10px;padding:3px 10px;border-radius:20px;border:1px solid #00b4d8}}
.kpi-row{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px}}
.kpi{{flex:1;min-width:120px;background:#0f3460;border-left:3px solid #00b4d8;padding:12px 16px;border-radius:6px}}
.kpi .num{{font-size:26px;font-weight:bold;color:#00b4d8}}
.kpi .lbl{{font-size:11px;color:#aaa;margin-top:2px}}
.kpi.green{{border-color:#06d6a0}}.kpi.green .num{{color:#06d6a0}}
.kpi.red{{border-color:#ef476f}}.kpi.red .num{{color:#ef476f}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#0f3460;color:#00b4d8;padding:10px 12px;text-align:left;font-size:12px}}
td{{padding:10px 12px;border-bottom:1px solid #0f3460;color:#e0e0e0}}
tr:hover td{{background:#0f3460}}
.footer{{text-align:center;color:#555;font-size:11px;padding:16px 30px;border-top:1px solid #0f3460}}
</style></head>
<body><div class='wrap'>
<div class='header'>
  <h1>Balanco por Vendedor ({periodo_inicio} ATE {periodo_fim})</h1>
  <p>Gerado as {hora_str} em {data_str} | RD Station CRM | MAXIFORCE</p>
</div>
<div class='body'>

<div class='section-title'>
  INDICADORES COMERCIAL (ultimos 3 meses)
</div>
<div class='kpi-row'>
  <div class='kpi'><div class='num'>{total_criadas_com}</div><div class='lbl'>Criadas</div></div>
  <div class='kpi'><div class='num'>{total_andamento}</div><div class='lbl'>Em Andamento</div></div>
  <div class='kpi green'><div class='num'>{total_ganhos_com}</div><div class='lbl'>Ganhos</div></div>
  <div class='kpi red'><div class='num'>{total_perdidos_com}</div><div class='lbl'>Perdidos</div></div>
  <div class='kpi green'><div class='num'>{_moeda(total_valor_com)}</div><div class='lbl'>Valor Ganhos</div></div>
</div>

<div class='section-title'>
  BALANCO POR VENDEDOR
  <span class='badge'>PYRAMID CLIENTES SEMANAL</span>
</div>
<table>
<tr><th>Vendedor</th><th>Criadas</th><th>Em Andamento</th><th>Finalizadas</th></tr>
{todas_rows}
</table>

</div>
<div class='footer'>Relatorio automatico - RD Station CRM | MAXIFORCE</div>
</div></body></html>"""
