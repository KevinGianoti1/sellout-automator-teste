
import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import openai
import os
from dotenv import load_dotenv

# Carrega variÃ¡veis do .env, inclusive OPENAI_API_KEY
load_dotenv()

# ğŸ“ Caminhos
input_path = Path("data/input")
output_path = Path("data/output")
output_path.mkdir(exist_ok=True)

# ğŸ“¥ Localiza o primeiro arquivo .xlsx
arquivo = next(input_path.glob("*.xlsx"), None)
if not arquivo:
    raise FileNotFoundError("Nenhum arquivo Excel encontrado em data/input")

print(f"ğŸ” Lendo arquivo: {arquivo.name}")
df = pd.read_excel(arquivo)

# ğŸ§® Calcula o total e extrai o ano da data de emissÃ£o
df["Total Calculado"] = df["Qtde"] * df["Valor Unit"]
df["Ano"] = pd.to_datetime(df["EmissÃ£o"]).dt.year

# ğŸ§‘â€ğŸ¤â€ğŸ§‘ Separar por representante
col_representante = "Repre"
representantes = df[col_representante].dropna().unique()

def gerar_grafico(df_rep, nome_arquivo_grafico):
    # Exemplo: Vendas por mÃªs
    df_rep['MÃªs'] = pd.to_datetime(df_rep['EmissÃ£o']).dt.month
    vendas_mes = df_rep.groupby('MÃªs')["Total Calculado"].sum()
    plt.figure(figsize=(8, 4))
    vendas_mes.plot(kind='bar')
    plt.title("Total de Vendas por MÃªs")
    plt.ylabel("R$ Vendido")
    plt.xlabel("MÃªs")
    plt.tight_layout()
    plt.savefig(nome_arquivo_grafico)
    plt.close()

def gerar_analise_gpt(df_rep):
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        return "Chave OpenAI nÃ£o encontrada. Crie um arquivo .env com OPENAI_API_KEY=sua_chave"
    openai.api_key = openai_api_key
    resumo = df_rep.describe().to_string()
    prompt = (
        f"Analise os resultados de vendas abaixo e dÃª um resumo dos principais pontos, tendÃªncias e sugestÃµes de melhoria:\n{resumo}"
    )
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}]
        )
        analise = response['choices'][0]['message']['content']
    except Exception as e:
        analise = f"Erro ao gerar anÃ¡lise com o ChatGPT: {e}"
    return analise

for rep in representantes:
    df_rep = df[df[col_representante] == rep].copy()

    # Organiza colunas na ordem desejada
    colunas_finais = [
        "Ano", "EmissÃ£o", "Notas", "Cliente", "CÃ³digo", "DescriÃ§Ã£o",
        "Qtde", "Valor Unit", "Total Calculado", "UF", "Empresa"
    ]
    df_saida = df_rep[colunas_finais].sort_values(by=["Ano", "EmissÃ£o"])

    # Salva o arquivo Excel bÃ¡sico
    nome_arquivo = f"Sell Out 2.0 - {rep}.xlsx"
    caminho_saida = output_path / nome_arquivo
    df_saida.to_excel(caminho_saida, index=False)

    # Gera grÃ¡fico e insere no Excel
    nome_arquivo_grafico = output_path / f"grafico_{rep}.png"
    gerar_grafico(df_rep, nome_arquivo_grafico)
    wb = load_workbook(caminho_saida)
    ws = wb.active
    img = Image(str(nome_arquivo_grafico))
    ws.add_image(img, 'L2')  # posiÃ§Ã£o do grÃ¡fico
    wb.save(caminho_saida)
    os.remove(nome_arquivo_grafico)

    # Gera anÃ¡lise com ChatGPT e insere em uma aba nova
    analise_gpt = gerar_analise_gpt(df_rep)
    ws_analise = wb.create_sheet("AnÃ¡lise GPT")
    ws_analise["A1"] = "AnÃ¡lise automÃ¡tica do ChatGPT:"
    for i, linha in enumerate(analise_gpt.split('\n'), 2):
        ws_analise[f"A{i}"] = linha
    wb.save(caminho_saida)

    print(f"âœ… Gerado: {nome_arquivo} com grÃ¡fico e anÃ¡lise GPT")

print("\nğŸ‰ RelatÃ³rios por representante gerados com sucesso!")
